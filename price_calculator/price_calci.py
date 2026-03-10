import pandas as pd
import io
import os
from datetime import datetime, timedelta
import openpyxl
# --- PLACEHOLDERS: Replace these with your actual download logic ---
def download_file(url, local_path):
    print(f"Downloading from {url} to {local_path}...")
    # Add your actual 'requests' or 'browser' automation code here
    pass

# --- 1. JEPX PROCESSING ---
def process_jepx(file_path):
    # Use encoding='utf-8-sig' to handle Japanese characters correctly
    df = pd.read_csv(file_path, encoding='utf-8-sig')
    
    # Filter for only "Non-FIT (Renewable Specified)"
    jepx_clean = df[df['商品'] == '非FIT(再エネ指定)'].copy()
    
    # Get the latest price based on auction date
    jepx_clean['約定日'] = pd.to_datetime(jepx_clean['約定日'])
    jepx_clean=jepx_clean.sort_values(by='約定日',ascending=True)

    return jepx_clean.values.tolist()

# --- 2. OCCTO PROCESSING ---
def process_occto(file_path):
    df = pd.read_csv(file_path, encoding='utf-8-sig')
    
    # Filter for Solar (太陽光)
    solar_df = df[df['電源種別'] == '太陽光'].copy()
    
    # Apply Reference Price Formula: Prev Annual Avg + Curr Month Avg - Prev Month Avg
    solar_df['参照価格'] = (
        solar_df['前年度平均価格'] + 
        solar_df['当年度月間平均価格'] - 
        solar_df['前年度月間平均価格']
    )

    # Extract year from filename (e.g., 'FY2025_sansyo_kakaku.csv' -> 2025)
    # This assumes the year is always in the filename
    filename = os.path.basename(file_path)
    year_from_file = int(''.join(filter(str.isdigit, filename)))
    solar_df['year'] = year_from_file
    
    # Parse month for the Excel mapping
    solar_df['date'] = pd.to_datetime(solar_df['年月'], format='%Y/%m')
    solar_df['month'] = solar_df['date'].dt.month
    # Example: Print latest calculation for a specific region (e.g., '中部')
    #latest_chubu = solar_df[solar_df['エリア'] == '中部'].sort_values(by='年月', ascending=False).iloc[0]
    
    #print(f"Latest Chubu Solar Reference Price: {latest_chubu['参照価格']} JPY/kWh")
    return solar_df

def find_year_row(sheet,year):
    for row in range(1,sheet.max_row + 1):
        cell_val=str(sheet.cell(row=row,column=1).value)
        if str(year)+"年度" in cell_val:
            return row
        return None
    
def find_region_row(sheet,region,start_row):
    #check region match in csv and excel rows
    for row in range(start_row+1,start_row+15):
        if sheet.cell(row=row,column=1).value==region:
            return row
    return None

def update_occto_excel(occto_df,excel_path):    
    wb=openpyxl.load_workbook(excel_path)
    sheet=wb['算出根拠']

    for year in occto_df['year'].unique():
        start_row = find_year_row(sheet,year)

        # --- Create new table if year not found ---
        if not start_row:
            prev_year_row=find_year_row(sheet,year-1)
            if prev_year_row:
                start_row=prev_year_row+12
            else:
                start_row=sheet.max_row+2
            sheet.cell(row=start_row,column=1,value=f"{year}年度")
            headers= ["4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月", "1月", "2月", "3月", "年度平均"]
            regions = ["北海道", "東北", "東京", "中部", "北陸", "関西", "中国", "四国", "九州", "沖縄"]
            for i,h in enumerate(headers):
                sheet.cell(row=start_row,column=i+1,value=h)
            for i,reg in enumerate(regions):
                sheet.cell(row=start_row+i+1,column=1,value=reg)
            print(f"Created table for {year} at row {start_row}")

        #map months
        col_map={}
        avg_col=None
        for col in range(2,15):
            head=str(sheet.cell(row=start_row,column=col).value)
            if '月' in head:
                m=int(head.replace('月','').strip())
                col_map[m]=col
            elif '年度平均' in head:
                avg_col = col

        #find empty month
        region_first=find_region_row(sheet,'北海道',start_row)
        target_col=None

        for m in col_map.keys():
            col=col_map[m]
            if sheet.cell(row=region_first,column=col).value is None:
                target_col=col
                print(f"Found 1st empty column! {m}月 (Column {col})")
                break
        
        #fill all regions of the empty column
        if target_col:
            year_data=occto_df[occto_df['year']==year]
            
            for m,col in col_map.items():
                if col>=target_col:
                    for _,row in year_data.iterrows():
                        if row['month']==m:
                            target_row=find_region_row(sheet,row['エリア'],start_row)
                            if target_row:
                                sheet.cell(row=target_row,column=col,value=row['参照価格'])
                                print(f"Filled {row['エリア']} {m}月 (Col {col})")
        # If Month 3 is filled, write the AVERAGE formula
        if region_first and avg_col:
            march_col=col_map.get(3)
            if march_col and sheet.cell(row=region_first,column=march_col).value is not None:
                for row_idx in range(region_first,region_first+10):
                    first_m_col = col_map.get(4) # 4月 start
                    last_m_col = col_map.get(3)  # 3月 end
                    formula=f"=AVERAGE({sheet.cell(row=row_idx, column=first_m_col).coordinate}:{sheet.cell(row=row_idx, column=last_m_col).coordinate})"
                    sheet.cell(row=row_idx,column=avg_col,value=formula)
                    print(f"Applied average formula to row {row_idx}")

    wb.save(excel_path)

def update_jepx_excel(jepx_df_row,excel_path):
    wb=openpyxl.load_workbook(excel_path)
    sheet=wb['算出根拠']
    start_col=16 #starts at column P

    # 1. FIND ANCHOR: The last row containing "非FIT(再エネ指定)" in the Product Column (Column R, which is index 18)
    header=None
    for row in range(1,sheet.max_row+1):
        if sheet.cell(row=row,column=start_col+3).value=='約定日':
            header=row
            break
    if not header:
        print("Error: Could not find JEPX header row in")
        return

    # COLLECT EXISTING DATES
    existing_date=set()
    curr_row=header+1
    while sheet.cell(row=curr_row,column=start_col).value is not None:
        val=sheet.cell(row=curr_row,column=start_col+3).value
        if val:
            existing_date.add(pd.to_datetime(val))
        curr_row+=1

    # write new data
    for data_row in jepx_df_row:
        csv_date=pd.to_datetime(data_row[3])

        if csv_date not in existing_date:
            for i,val in enumerate(data_row):
                sheet.cell(row=curr_row,column=start_col+i,value=val)
            # 1. Define the start and end of the 4-row window
            # If we are writing to row 33, window is 30:33 (33-3 = 30)
            end_window = curr_row
            start_window = max(header + 1, curr_row - 3) 

            # 2. Get column letters dynamically (Column 19 is T, Column 20 is U)
            # Ensure these point to the correct indices in your Excel sheet
            vol_col = start_col + 4 # Column T
            price_col = start_col + 5 # Column U

            vol_letter = sheet.cell(row=header+1, column=vol_col).column_letter
            price_letter = sheet.cell(row=header+1, column=price_col).column_letter

            # 3. Construct the formula
            formula = (
                f"=SUMPRODUCT({price_letter}{start_window}:{price_letter}{end_window}, "
                f"{vol_letter}{start_window}:{vol_letter}{end_window}) / "
                f"SUM({vol_letter}{start_window}:{vol_letter}{end_window})"
            )

            # 4. Write the formula into the Weighted Average column
            avg_col_idx = start_col + 12 # Column AD (example)
            sheet.cell(row=curr_row, column=avg_col_idx, value=formula)
            
            print(f"Added record and formula at row {curr_row}")
            curr_row += 1
            existing_date.add(csv_date)

    wb.save(excel_path)

# --- 3. MAIN LOGIC (THE "CHECK-IF-STALE" FLOW) ---
def run_pipeline():
    jepx_file = 'nf_summary_2025.csv'
    occto_file = 'FY2025_sansyo_kakaku.csv'
    
    # Logic: If stale/missing -> Download -> Process
    # (Placeholder logic below)
    if not os.path.exists(jepx_file): # Or check time diff > 90 days
        download_file("https://jepx.url", jepx_file)
        
    if not os.path.exists(occto_file): # Or check last month > 90 days
        download_file("https://occto.url", occto_file)
    price_calci="D:\VS_CODE\Infiswift\簡易シミュレーター.xlsx"
    # Process the files
    jepx_val = process_jepx(jepx_file)
    occto_data = process_occto(occto_file)
    update_occto_excel(occto_data,price_calci)
    update_jepx_excel(jepx_val,price_calci)
if __name__ == "__main__":
    run_pipeline()