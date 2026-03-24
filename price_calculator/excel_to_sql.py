import pandas as pd
import openpyxl
from sqlalchemy import create_engine
import win32com.client as win32
import os

# --- 1. Database Connection ---
# Replace with your actual MySQL credentials
engine = create_engine("mysql+mysqlconnector://root:params1812@localhost:3306/priceCalci")

def find_anchor(sheet, keyword):
    """Scans the entire sheet to find the exact row and col of a specific text."""
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and keyword in cell.value:
                return cell.row, cell.column
    return None, None

def extract_exact_table(sheet, anchor_text, header_offset, num_cols):
    """
    Extracts a table EXACTLY as it appears in Excel (Wide Format).
    - header_offset: How many rows below the anchor the column headers start.
    - num_cols: The exact number of columns to read left-to-right.
    """
    r, c = find_anchor(sheet, anchor_text)
    if not r:
        print(f"❌ Anchor '{anchor_text}' not found.")
        return None

    # 1. Read Headers
    header_r = r + header_offset
    headers = []
    for i in range(num_cols):
        val = sheet.cell(row=header_r, column=c + i).value
        # Clean up headers (remove newlines/spaces) or assign default if empty
        clean_header = str(val).strip().replace("\n", "") if val is not None else f"Column_{i+1}"
        headers.append(clean_header)

    # 2. Read Data Rows
    data = []
    curr_r = header_r + 1
    
    # Keep reading downwards until the FIRST column of the row is empty
    while sheet.cell(row=curr_r, column=c).value is not None:
        row_data = {}
        for i in range(num_cols):
            row_data[headers[i]] = sheet.cell(row=curr_r, column=c + i).value
        data.append(row_data)
        curr_r += 1

    # 3. Convert to DataFrame
    df = pd.DataFrame(data)
    print(f"✅ Extracted '{anchor_text}': {len(df)} rows, {num_cols} columns.")
    return df
"""
def force_excel_calculation(excel_path):
    # Opens Excel invisibly, forces it to calculate all formulas, saves the cached values, and closes it.
    print("⚙️ Forcing Excel to calculate formulas...")
    abs_path = os.path.abspath(excel_path) # win32com requires absolute paths!
    
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False # Keep it hidden in the background
    excel.DisplayAlerts = False # Ignore popups
    
    try:
        wb = excel.Workbooks.Open(abs_path)
        wb.Save() # This forces Excel to calculate and save the cached values!
        wb.Close(SaveChanges=True)
        print("✅ Formulas calculated and saved.")
    except Exception as e:
        print(f"❌ Error refreshing Excel: {e}")
    finally:
        excel.Quit()"""

def migrate_exact_tables(excel_path):
    #force_excel_calculation(excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb['算出根拠'] # IMPORTANT: Update if your sheet name is different
    
    # --- Configuration for your 4 tables ---
    # Format: (Anchor Text, Header Row Offset, Number of Columns, MySQL Table Name)
    '''table_configs = [
        ("b.参照価格", 1, 7, "reference_price"),
        ("c.非化石価値相当額", 8, 13, "non_fossil"),
        ("d.バランシングコスト", 1, 3, "balancing_cost"),
        ("⑤PPA単価", 1, 2, "ppa_price")
    ]'''
    table_configs = [
        ("b.参照価格", 1, 7, "reference_price"),
        ("c.非化石価値相当額", 8, 13, "non_fossil")]

    for anchor, offset, cols, db_table_name in table_configs:
        df = extract_exact_table(sheet, anchor, offset, cols)
        
        if df is not None and not df.empty:
            # Show a preview of what we grabbed before sending to MySQL
            print(f"Preview of {db_table_name}:")
            print(df.head(3))
            print("-" * 40)
            
            # Write exactly as-is to MySQL
            df.to_sql(name=db_table_name, con=engine, if_exists='replace', index=False)
            print(f"🚀 Successfully written to MySQL table: {db_table_name}\n")

if __name__ == "__main__":
    migrate_exact_tables(r"D:\VS_CODE\Infiswift\簡易シミュレーター.xlsx")
