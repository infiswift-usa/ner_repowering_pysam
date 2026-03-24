"""
data_extractor.py

Run this ONCE (or on a schedule) whenever you get new JEPX / OCCTO CSV files.
 - Updates the Excel workbook (算出根拠 sheet) with new data
 - Uses Gemini AI to read the resulting tables and extract clean JSON config
 - Saves simulator_config.json for main_simulator.py to consume

Nothing is hardcoded.  All constants come from the live Excel sheet.
"""

import os
import json
import pandas as pd
import openpyxl
from datetime import datetime
from google import genai
from dotenv import load_dotenv

# ── Load API Key ──────────────────────────────────────────────────────────────
load_dotenv()
API_KEY = os.getenv("google_api_key") or os.getenv("GOOGLE_API_KEY")
client = genai.Client(api_key=API_KEY)

EXCEL_PATH = r"D:\VS_CODE\Infiswift\簡易シミュレーター.xlsx"
JSON_OUT   = "simulator_config.json"


# ══════════════════════════════════════════════════════════════════════════════
# 1.  JEPX  ── Update Excel sheet with new auction data
# ══════════════════════════════════════════════════════════════════════════════

def process_jepx_csv(file_path: str) -> list:
    try:
        df = pd.read_csv(file_path, encoding="shift-jis")
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, encoding="utf-8-sig")

    clean = df[df["商品"] == "非FIT(再エネ指定)"].copy()
    # CSV 約定日 format: YYYY/M/D  (e.g. "2025/8/28")
    # Excel stores dates as M/D/YYYY strings (e.g. "8/28/2025") — convert explicitly.
    parsed = pd.to_datetime(clean["約定日"], format="%Y/%m/%d")
    clean["約定日"] = parsed.apply(lambda d: f"{d.month}/{d.day}/{d.year}")
    clean = clean.sort_values("約定日", key=lambda s: pd.to_datetime(s, format="%m/%d/%Y"), ascending=True)
    return clean.values.tolist()


def update_jepx_excel(jepx_rows: list, wb: openpyxl.Workbook) -> None:
    sheet = wb["算出根拠"]
    start_col = 16  # Column P

    # Find header row (column S = start_col+3 contains "約定日")
    header_row = None
    for r in range(1, sheet.max_row + 1):
        if sheet.cell(r, start_col + 3).value == "約定日":
            header_row = r
            break
    if not header_row:
        print("ERROR: JEPX header row not found – skipping JEPX update.")
        return

    # Collect existing dates to avoid duplicates.
    # Normalize to "M/D/YYYY" string regardless of how openpyxl returns the value
    def _to_date_str(val) -> str | None:
        if val is None:
            return None
        if isinstance(val, datetime):
            return f"{val.month}/{val.day}/{val.year}"
        try:
            d = pd.to_datetime(str(val), dayfirst=False)
            return f"{d.month}/{d.day}/{d.year}"
        except Exception:
            return str(val)  # keep as-is if unparseable

    existing: set[str] = set()
    cur = header_row + 1
    while sheet.cell(cur, start_col).value is not None:
        ds = _to_date_str(sheet.cell(cur, start_col + 3).value)
        if ds:
            existing.add(ds)
        cur += 1
    
    vol_col   = start_col + 4   # 約定総量
    price_col = start_col + 5   # 約定価格
    avg_col   = start_col + 12  # 加重平均値

    # Write new rows
    for row_data in jepx_rows:
        # row_data[3] is already a M/D/YYYY string after process_jepx_csv
        date_str = str(row_data[3]) 
        if date_str in existing:
            continue

        for i, v in enumerate(row_data):
            sheet.cell(cur, start_col + i, value=v)

        # Rolling 4-auction volume-weighted average formula
        end_w   = cur
        start_w = max(header_row + 1, cur - 3)
        v_col   = sheet.cell(header_row + 1, vol_col).column_letter
        p_col   = sheet.cell(header_row + 1, price_col).column_letter
        formula = (
            f"=SUMPRODUCT({p_col}{start_w}:{p_col}{end_w},"
            f"{v_col}{start_w}:{v_col}{end_w})"
            f"/SUM({v_col}{start_w}:{v_col}{end_w})"
        )
        sheet.cell(cur, avg_col, value=formula)

        print(f"  [JEPX] Added row {cur}: {date_str}")
        existing.add(date_str)
        cur += 1

# ══════════════════════════════════════════════════════════════════════════════
# 2.  OCCTO ── Update Excel sheet with new 参照価格 data
# ══════════════════════════════════════════════════════════════════════════════

def process_occto_csv(file_path: str) -> pd.DataFrame:
    try:
        df = pd.read_csv(file_path, encoding="shift-jis")
    except UnicodeDecodeError:
        df = pd.read_csv(file_path, encoding="utf-8-sig")

    solar = df[df["電源種別"] == "太陽光"].copy()
    solar["参照価格"] = (
        solar["前年度平均価格"]
        + solar["当年度月間平均価格"]
        - solar["前年度月間平均価格"]
    )

    digits = "".join(filter(str.isdigit, os.path.basename(file_path)))
    solar["year"]  = int(digits) if digits else datetime.now().year
    solar["date"]  = pd.to_datetime(solar["年月"], format="%Y/%m")
    solar["month"] = solar["date"].dt.month
    return solar

def _find_row_containing(sheet, col: int, text: str, stop: int = 300) -> int | None:
    for r in range(1, stop):
        v = sheet.cell(r, col).value
        if v and str(text) in str(v):
            return r
    return None

def update_occto_excel(occto_df: pd.DataFrame, wb: openpyxl.Workbook) -> None:
    sheet = wb["算出根拠"]
    regions = ["北海道", "東北", "東京", "中部", "北陸", "関西", "中国", "四国", "九州", "沖縄"]

    for year in occto_df["year"].unique():
        start_row = _find_row_containing(sheet, 1, f"{year}年度")

        # year is retreived from filename. if we dont find the row with year - it means we need to create new table
        if start_row is None:
            prev = _find_row_containing(sheet, 1, f"{year-1}年度")
            start_row = (prev + 12) if prev else (sheet.max_row + 2)
            sheet.cell(start_row, 1, value=f"{year}年度")
            for i, h in enumerate(["4月","5月","6月","7月","8月","9月","10月","11月","12月","1月","2月","3月","年度平均"]):
                sheet.cell(start_row, i + 2, value=h)
            for i, reg in enumerate(regions):
                sheet.cell(start_row + i + 1, 1, value=reg)
            print(f"  [OCCTO] Created table for {year} at row {start_row}")

        # Build month→column and avg column maps
        col_map, avg_col = {}, None
        for c in range(2, 16):
            h = str(sheet.cell(start_row, c).value or "")
            if "月" in h:
                col_map[int(h.replace("月","").strip())] = c
            elif "年度平均" in h:
                avg_col = c
        
        region_row_map={}
        for r in range(start_row+1,start_row+15):
            val=sheet.cell(r,1).value
            if val in regions:
                region_row_map[val]=r

        #fill in table . if year doesnt exist above snippet adds structure.if it exists, below snippet fills in cells
        # Find first empty month column
        first_reg_row = start_row + 1
        target_col = None
        if first_reg_row:
            for m, c in col_map.items():
                if sheet.cell(first_reg_row, c).value is None:
                    target_col = c
                    break

        # Fill data
        if target_col:
            year_data = occto_df[occto_df["year"] == year]
            for m, c in col_map.items():
                if c >= target_col:
                    month_data=year_data[year_data["month"]==m]

                    for _, row in year_data.iterrows():
                        if row["month"] == m:
                            reg_row=region_row_map.get(row["エリア"])
                            if reg_row:
                                sheet.cell(reg_row, c, value=row["参照価格"])

        # If March (month 3) is now filled → write annual average formula
        march_col = col_map.get(3)
        if first_reg_row and avg_col and march_col:
            if sheet.cell(first_reg_row, march_col).value is not None:
                apr_col = col_map.get(4)
                for ri in range(first_reg_row, first_reg_row + 10):
                    c1 = sheet.cell(ri, apr_col).coordinate
                    c2 = sheet.cell(ri, march_col).coordinate
                    sheet.cell(ri, avg_col, value=f"=AVERAGE({c1}:{c2})")

    print("  [OCCTO] Excel update done.")


# ══════════════════════════════════════════════════════════════════════════════
# 3.  AI EXTRACTION ── Use Gemini to read tables from the workbook
# ══════════════════════════════════════════════════════════════════════════════

'''def _sheet_to_text(wb: openpyxl.Workbook, sheet_name: str) -> str:
    """Dump the sheet to a compact CSV-like text for the LLM."""
    sheet = wb[sheet_name]
    lines = []
    for r in range(1, sheet.max_row + 1):
        cells = []
        has_data = False
        for c in range(1, sheet.max_column + 1):
            v = sheet.cell(r, c).value
            cells.append("" if v is None else str(v))
            if v is not None:
                has_data = True
        if has_data:
            lines.append(",".join(cells))
    
    return "\n".join(lines)'''

def _sheet_to_text(wb:openpyxl.Workbook,sheet_name:str)->str:
    #dump sheet into compact CSV like text for LLM using pandas
    df=pd.DataFrame(wb[sheet_name].values)
    df=df.dropna(how="all",axis=0).dropna(how="all",axis=1) #only drops row/column that are 100% empty
    return df.to_csv(index=False,header=False)

def extract_config_with_ai(wb: openpyxl.Workbook) -> dict:
    """
    Feed the 算出根拠 sheet text to Gemini and ask it to extract:
      - b.参照価格 per region (2020-2024 average excluding 2022)
      - c.非化石価値相当額 = 加重平均値 from the LAST data row of the JEPX table
      - d.バランシングコスト list (20 values, one per year 2026-2045)
      - ⑤PPA単価
    Returns a clean dict.
    """
    print("  [AI] Dumping sheet to text for Gemini…")
    sheet_text = _sheet_to_text(wb, "算出根拠")
    prompt = f"""
    You are a data extraction assistant. Below is a raw CSV dump of a Japanese Excel worksheet called 「算出根拠」.

    Your task: Return a single valid JSON object with exactly these keys:

    1."reference_prices":dict mapping each Japanese region name to its float value from the column header "2020-2024平均(2022年度除く)". (From the b.参照価格 table)
    2. "non_fossil_value": float - the final numeric value (加重平均値) from the VERY LAST data row of the JEPX table under "c.非化石価値相当額".
    3. "balancing_costs":list if floates - the numeric values from the d.バランシングコスト column.
    4. "ppa_prices": dict mapping each region name to its float value STRICTLY UNDER the table "⑤PPA単価" (e.g., {{"北海道": 14.0, "東京": 14.5}}).

    Return ONLY the JSON object. No explanation, no markdown fences. Do not include markdown formatting like ```json.

    Sheet data:
    {sheet_text}
    """

    resp = client.models.generate_content(model="gemini-3-flash-preview",contents=prompt)
    raw = resp.text.strip()

    # Strip any accidental markdown fences
    if raw.startswith("```"):
        raw = "\n".join(
            line for line in raw.splitlines()
            if not line.strip().startswith("```")
        ).strip()

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"  [AI] JSON parse failed: {e}\n  Raw response:\n{raw[:500]}")
        raise

    print(f"  [AI] Extracted reference_prices: {list(data.get('reference_prices',{}).keys())}")
    print(f"  [AI] non_fossil_value = {data.get('non_fossil_value')}")
    print(f"  [AI] balancing_costs ({len(data.get('balancing_costs',[]))} values)")
    print(f"  [AI] Extracted ppa_prices: {list(data.get('ppa_prices',{}).keys())}")
    return data


# ══════════════════════════════════════════════════════════════════════════════
# 4.  MAIN RUNNER
# ══════════════════════════════════════════════════════════════════════════════

def run_extractor(
    jepx_csv: str  = "nf_summary_2025.csv",
    occto_csv: str = "FY2025_sansyo_kakaku.csv",
    excel_path: str = EXCEL_PATH,
    json_out: str   = JSON_OUT,
):
    print("=" * 60)
    print("DATA EXTRACTOR")
    print("=" * 60)

    # ── Load workbook (with formulas so we can write)
    wb = openpyxl.load_workbook(excel_path)
    update_needed=False
    # ── Update JEPX section
    if update_needed:
        if os.path.exists(jepx_csv):
            print(f"\n>>> Updating JEPX from {jepx_csv}")
            rows = process_jepx_csv(jepx_csv)
            update_jepx_excel(rows, wb)
        else:
            print(f"[SKIP] {jepx_csv} not found — JEPX section not updated.")

        # ── Update OCCTO section
        if os.path.exists(occto_csv):
            print(f"\n>>> Updating OCCTO from {occto_csv}")
            occto_df = process_occto_csv(occto_csv)
            update_occto_excel(occto_df, wb)
        else:
            print(f"[SKIP] {occto_csv} not found — OCCTO section not updated.")

    # ── Save the updated workbook so formulas are stored
    try:
        if update_needed:
            wb.save(excel_path)
            print(f"\n>>> Saved updated workbook: {excel_path}")
        else:
            print("update_needed is turned off -- continuing without updating data")
    except PermissionError:
        print(
            "\n[ERROR] Cannot save Excel file — it is currently open in Excel.\n"
            "  → Please CLOSE the file in Excel and run this script again."
        )
        return

    # ── Reload with data_only=True so openpyxl reads the cached formula results
    # NOTE: cached values are from the last time Excel opened and saved the file.
    # For formula results to be fresh, open and save the file in Excel once, THEN run this.
    # Alternatively, Gemini reads the raw data directly from the formula-version sheet (more reliable).
    wb_for_ai = openpyxl.load_workbook(excel_path, data_only=True)

    # ── Use Gemini AI to extract the config
    print("\n>>> Asking Gemini AI to extract config from sheet…")
    try:
        ai_data = extract_config_with_ai(wb_for_ai)
    except Exception as e:
        print(f"[ERROR] AI extraction failed: {e}")
        print("Cannot build config. Exiting.")
        return

    # ── Assemble final config
    config = {
        "reference_prices": ai_data["reference_prices"],
        "non_fossil_value": ai_data["non_fossil_value"],
        "balancing_costs":  ai_data["balancing_costs"],
        "ppa_prices":       ai_data["ppa_prices"],
    }

    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

    print(f"\n>>> Config saved to: {json_out}")
    print(json.dumps(config, indent=2, ensure_ascii=False))
    return config


if __name__ == "__main__":
    run_extractor()
